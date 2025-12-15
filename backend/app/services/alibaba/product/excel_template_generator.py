from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.comments import Comment


class ExcelTemplateGenerator:
    """
    Schema → Excel Template
    """
    BASIC_PATH_WHITELIST = {
        "productTitle",
        "saleType",
        "scPrice",
        "priceUnit",
        "minOrderQuantity",
        "superText",
        "catId",
    }

    def __init__(self, schema_json: list[dict]):
        self.schema = schema_json
        self.fields = []
        self._flatten(schema_json)

    def _flatten(self, fields):
        for f in fields:
            if f.get("type") not in ("label",):
                self.fields.append(f)
            if f.get("children"):
                self._flatten(f["children"])

    # ================= Public =================

    def generate(self, filename="products_template.xlsx"):
        wb = Workbook()

        self._gen_readme(wb)
        self._gen_basic(wb)
        self._gen_attributes(wb)
        self._gen_images(wb)
        self._gen_package(wb)
        self._gen_sku(wb)

        wb.remove(wb.active)
        wb.save(filename)

    # ================= README =================

    def _gen_readme(self, wb):
        ws = wb.create_sheet("README")
        ws["A1"] = "填写说明（由 Schema 自动生成）"
        ws["A1"].font = Font(bold=True)

        rules = [
            "1. 每一行 = 一个产品（以 Product name / productTitle 关联）",
            "2. multiCheck 字段使用 | 分隔",
            "3. 支持 Other:xxx（当 Schema 允许自定义值）",
            "4. 不要修改列名",
            "5. 本模板完全由 Schema 自动生成",
        ]

        for i, r in enumerate(rules, start=3):
            ws[f"A{i}"] = r

    # ================= BASIC =================

    def _gen_basic(self, wb):
        ws = wb.create_sheet("product_basic")

        for f in self.fields:
            if f["path"] in self.BASIC_PATH_WHITELIST:
                self._append_column(ws, f)

    # ================= ATTRIBUTES =================

    def _gen_attributes(self, wb):
        ws = wb.create_sheet("product_attributes")

        for f in self.fields:
            if f["path"].startswith("icbuCatProp."):
                if self._is_fillable(f):
                    self._append_column(ws, f)

    # ================= IMAGES =================

    def _gen_images(self, wb):
        ws = wb.create_sheet("product_images")
        ws.append(["productTitle", "imageUrl"])

    # ================= PACKAGE =================

    def _gen_package(self, wb):
        ws = wb.create_sheet("product_package")

        for f in self.fields:
            if f["path"].startswith("pkgMeasure") or f["path"] == "pkgWeight":
                if self._is_fillable(f):
                    self._append_column(ws, f)

    # ================= SKU =================

    def _gen_sku(self, wb):
        ws = wb.create_sheet("product_sku")

        for f in self.fields:
            if f["path"].startswith("saleProp."):
                if self._is_fillable(f):
                    self._append_column(ws, f)

        ws.append(["price", "stock"])

    # ================= Helpers =================

    def _is_fillable(self, field: dict) -> bool:
        """
        是否值得生成 Excel 列
        """
        return (
            field["type"] in ("input", "singleCheck", "multiCheck")
            and (
                field.get("required")
                or field.get("options")
                or field["path"] in self.BASIC_PATH_WHITELIST
            )
        )

    def _append_column(self, ws, field):
        cell = ws.cell(row=1, column=ws.max_column + 1)
        cell.value = field["name"] or field["id"]
        cell.font = Font(bold=True)
        cell.comment = self._field_comment(field)

    def _field_comment(self, field):
        lines = []

        lines.append(f"path: {field['path']}")

        if field.get("required"):
            lines.append("必填")

        if field["type"] == "multiCheck":
            lines.append("多选：使用 | 分隔")

        if field.get("options"):
            opts = [o.get("displayName") for o in field["options"]]
            lines.append("可选值：" + ", ".join(opts[:10]))

        return Comment("\n".join(lines), "SchemaBot")
